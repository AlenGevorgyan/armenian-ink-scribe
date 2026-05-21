"""
File export utilities for Araks OCR.

Generates downloadable files from processed OCR text:
  - XLSX / CSV  for tabular data
  - PDF / TXT   for plain text
"""

import csv
import io
import os
import logging
from typing import List

logger = logging.getLogger(__name__)

# --------------- font path for PDF generation with Armenian glyphs ---------------
# We look in several places so it works both locally and in Docker.
_FONT_SEARCH_PATHS = [
    os.path.join(os.path.dirname(__file__), "fonts", "NotoSansArmenian-Regular.ttf"),
    "/usr/share/fonts/truetype/noto/NotoSansArmenian-Regular.ttf",
    os.path.join(os.path.dirname(__file__), "NotoSansArmenian-Regular.ttf"),
]

_ARMENIAN_FONT_PATH: str = ""
for _p in _FONT_SEARCH_PATHS:
    if os.path.isfile(_p):
        _ARMENIAN_FONT_PATH = _p
        break

if _ARMENIAN_FONT_PATH:
    logger.info("Armenian font found at %s", _ARMENIAN_FONT_PATH)
else:
    logger.warning(
        "Armenian font NOT found — PDF export will fall back to built-in font "
        "(Armenian glyphs may not render). Place NotoSansArmenian-Regular.ttf "
        "in the 'fonts/' directory or install noto fonts system-wide."
    )


# ====================================================================== XLSX

def export_table_xlsx(rows: List[List[str]]) -> bytes:
    """Create an Excel (.xlsx) file from a list of rows.

    The first row is treated as headers and styled with a bold font.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Table"

    # Header style
    header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

    # Auto-size columns (approximate)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ====================================================================== CSV

def export_table_csv(rows: List[List[str]]) -> bytes:
    """Create a UTF-8 CSV from a list of rows."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat


# ====================================================================== PDF

def export_text_pdf(text: str, title: str = "OCR Result") -> bytes:
    """Create a PDF with the corrected Armenian text.

    Uses reportlab with Noto Sans Armenian when available.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT

    buf = io.BytesIO()

    # Register Armenian font if available
    font_name = "Helvetica"
    if _ARMENIAN_FONT_PATH:
        try:
            pdfmetrics.registerFont(TTFont("NotoSansArmenian", _ARMENIAN_FONT_PATH))
            font_name = "NotoSansArmenian"
        except Exception as exc:
            logger.warning("Failed to register Armenian font: %s", exc)

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=25 * mm,
        bottomMargin=25 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "OCRTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=18,
        spaceAfter=12,
    )

    body_style = ParagraphStyle(
        "OCRBody",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=12,
        leading=18,
        alignment=TA_LEFT,
        spaceAfter=8,
    )

    story = []
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10 * mm))

    # Split text into paragraphs and add them
    for para in text.split("\n"):
        para = para.strip()
        if not para:
            story.append(Spacer(1, 4 * mm))
        else:
            # Escape XML special characters for reportlab
            safe = (
                para.replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
            )
            story.append(Paragraph(safe, body_style))

    doc.build(story)
    return buf.getvalue()


# ====================================================================== TXT

def export_text_txt(text: str) -> bytes:
    """Return UTF-8 encoded plain text."""
    return text.encode("utf-8")
